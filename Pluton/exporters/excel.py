"""
Export des fiches Excel — une ligne par entreprise, colorisée par niveau de chaleur.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# (libellé colonne, clé dans fiche_excel)
COLUMNS = [
    ("Entreprise", "entreprise_nom"),
    ("SIREN", "siren"),
    ("Code APE", "code_ape"),
    ("Ville", "ville"),
    ("Région", "region"),
    ("CA (€)", "ca_euros"),
    ("Effectifs", "effectifs"),
    ("Évolution effectifs", "evolution_effectifs"),
    ("Nb agences", "nb_agences"),
    ("Groupe / filiales", "groupe_filiales"),
    ("CRM détecté", "stack_crm"),
    ("Autres outils", "stack_autres"),
    ("Site web", "site_web"),
    ("Prénom", "decideur_prenom"),
    ("Nom", "decideur_nom"),
    ("Poste", "decideur_poste"),
    ("Email", "decideur_email"),
    ("Téléphone", "decideur_telephone"),
    ("LinkedIn", "decideur_linkedin"),
    ("Ancienneté (mois)", "decideur_anciennete_mois"),
    ("Score", "score"),
    ("Chaleur", "niveau_chaleur"),
    ("Priorité", "priorite"),
    ("Signaux détectés", "signaux_resume"),
    ("Accroche", "angle_accroche"),
    ("Objet email", "email_objet"),
    ("Corps email", "email_corps"),
    ("Date traitement", "date_traitement"),
]

COLOR_MAP = {
    "très chaud": "FF4444",
    "chaud": "FF9900",
    "tiède": "FFDD00",
    "froid": "AAAAAA",
}

WIDTH_MAP = {
    "Entreprise": 30, "Signaux détectés": 40, "Accroche": 40,
    "Objet email": 40, "Corps email": 60, "Site web": 30,
    "LinkedIn": 30, "Email": 28, "Poste": 22, "Groupe / filiales": 24,
}


def export_to_excel(results: list, output_dir: str = ".") -> str:
    """Construit le fichier Excel à partir des dossiers produits par l'agent."""
    if not results:
        print("[Excel] Aucune fiche à exporter.")
        return ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects Pluton"

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = WIDTH_MAP.get(label, 18)
    ws.row_dimensions[1].height = 32

    # Tri par score décroissant.
    def _score(r):
        try:
            return float(r.get("fiche_excel", {}).get("score", 0) or 0)
        except (TypeError, ValueError):
            return 0
    results = sorted(results, key=_score, reverse=True)

    chaleur_col = next(i + 1 for i, (_, k) in enumerate(COLUMNS) if k == "niveau_chaleur")

    for row_idx, result in enumerate(results, start=2):
        fiche = result.get("fiche_excel", {})
        chaleur = str(fiche.get("niveau_chaleur", "froid")).lower()
        color = COLOR_MAP.get(chaleur, "FFFFFF")

        for col_idx, (_, key) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=fiche.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == chaleur_col:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="FFFFFF")
        ws.row_dimensions[row_idx].height = 70

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filepath = os.path.join(output_dir, f"prospects_pluton_{timestamp}.xlsx")
    wb.save(filepath)

    print(f"\n[Excel] Fichier exporté : {filepath} ({len(results)} entreprises)")
    return filepath
